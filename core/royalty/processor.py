# vcpmctool/core/royalty/processor.py
"""
Module xử lý file Excel cho tính nhuận bút
Đã sửa: Chỉ tính mức nhuận bút gia hạn khi có ngày gia hạn tương ứng
Thêm mới: Cột Link YouTube với timestamp ở cuối
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from typing import Dict, Tuple, Callable, Optional
from datetime import datetime
from dateutil.relativedelta import relativedelta

from .calculator import RoyaltyCalculator
from ..datefmt import parse_date, to_ddmmyyyy


class RoyaltyProcessor:
    """Xử lý file Excel với tính toán nhuận bút"""

    def __init__(self, royalty_dict: Dict[str, Tuple[int, int, int]]):
        self.calculator = RoyaltyCalculator(royalty_dict)
        self.errors = []

    def _create_youtube_link_with_timestamp(self, video_id: str, time_range: str) -> str:
        """
        Tạo link YouTube với timestamp từ ID và khoảng thời gian

        Args:
            video_id: YouTube video ID (11 ký tự)
            time_range: Chuỗi thời gian dạng "mm:ss - mm:ss" hoặc "hh:mm:ss - hh:mm:ss"

        Returns:
            Link YouTube với timestamp
        """
        if not video_id or pd.isna(video_id) or not time_range or pd.isna(time_range):
            return ""

        video_id = str(video_id).strip()
        time_range = str(time_range).strip()

        # Kiểm tra video ID hợp lệ (11 ký tự)
        if len(video_id) != 11:
            return ""

        # Kiểm tra có dấu '-' để phân tách khoảng thời gian
        if '-' not in time_range:
            return f"https://www.youtube.com/watch?v={video_id}"

        try:
            # Lấy thời gian bắt đầu (phần trước dấu '-')
            start_time = time_range.split('-')[0].strip()

            # Chuyển đổi thành giây
            parts = start_time.split(':')

            if len(parts) == 2:  # Format mm:ss
                minutes = int(parts[0])
                seconds = int(parts[1])
                total_seconds = minutes * 60 + seconds
            elif len(parts) == 3:  # Format hh:mm:ss
                hours = int(parts[0])
                minutes = int(parts[1])
                seconds = int(parts[2])
                total_seconds = hours * 3600 + minutes * 60 + seconds
            else:
                # Format không hợp lệ, trả về link không có timestamp
                return f"https://www.youtube.com/watch?v={video_id}"

            # Tạo link với timestamp
            return f"https://www.youtube.com/watch?v={video_id}&t={total_seconds}s"

        except (ValueError, IndexError):
            # Lỗi khi parse, trả về link không có timestamp
            return f"https://www.youtube.com/watch?v={video_id}"

    def process_file(
            self,
            input_path: str,
            output_path: str,
            progress_callback: Optional[Callable] = None,
            log_callback: Optional[Callable] = None
    ) -> Tuple[bool, str]:
        """
        Xử lý file Excel và tính nhuận bút
        Returns: (success, message)
        """
        try:
            if log_callback:
                log_callback("🔍 Đang kiểm tra file đầu vào...")
                
            # Đọc file Excel
            df = pd.read_excel(input_path, engine='openpyxl')

            if df.empty:
                return False, "❌ File Excel không có dữ liệu hoặc định dạng không đúng"
                
            if log_callback:
                log_callback(f"📊 Đã đọc {len(df)} dòng dữ liệu")

            # Xử lý từng dòng
            total_rows = len(df)
            processed_data = []
            
            if log_callback:
                log_callback("⚙️ Bắt đầu xử lý và tính toán nhuận bút...")

            for idx, row in df.iterrows():
                if progress_callback:
                    progress = ((idx + 1) / total_rows) * 100
                    progress_callback(progress)

                # +2 vì Excel bắt đầu từ 1 và có header
                processed_row = self._process_row(row, idx + 2)
                processed_data.append(processed_row)
                
            if log_callback:
                log_callback("💾 Đang tạo file Excel với định dạng...")

            # Tạo DataFrame mới với dữ liệu đã xử lý
            result_df = pd.DataFrame(processed_data)

            # THÊM CỘT LINK Ở CUỐI CÙNG
            # Di chuyển cột Link YouTube với timestamp vào cuối
            if 'Link YouTube Timestamp' in result_df.columns:
                # Lấy danh sách cột hiện tại trừ cột Link
                cols = [col for col in result_df.columns if col != 'Link YouTube Timestamp']
                # Thêm cột Link vào cuối
                cols.append('Link YouTube Timestamp')
                # Sắp xếp lại DataFrame
                result_df = result_df[cols]

            # Ghi ra file Excel với định dạng
            success = self._write_formatted_excel(result_df, output_path)

            if success:
                if log_callback:
                    log_callback("✅ Hoàn tất! Đã thêm cột Link YouTube với timestamp")
                    log_callback(f"📁 File kết quả: {output_path}")
                return True, f"✅ Xử lý thành công!\n\n📁 Kết quả đã được lưu tại:\n{output_path}\n\n🔗 File bao gồm:\n• Mức nhuận bút được tính tự động\n• Link YouTube với timestamp\n• Định dạng Excel chuyên nghiệp"
            else:
                return False, "❌ Lỗi khi ghi file Excel. Vui lòng kiểm tra:\n• File kết quả có đang mở không?\n• Có quyền ghi vào thư mục không?"

        except Exception as e:
            error_msg = str(e)
            if "Permission denied" in error_msg:
                return False, "❌ Lỗi quyền truy cập!\n\nVui lòng:\n• Đóng file Excel nếu đang mở\n• Chạy ứng dụng với quyền Administrator\n• Kiểm tra quyền ghi vào thư mục"
            elif "No such file" in error_msg:
                return False, "❌ Không tìm thấy file!\n\nVui lòng kiểm tra:\n• Đường dẫn file có đúng không?\n• File có bị di chuyển hoặc xóa không?"
            else:
                return False, f"❌ Lỗi không xác định:\n\n{error_msg}\n\nVui lòng liên hệ hỗ trợ kỹ thuật."

    def _process_row(self, row: pd.Series, excel_row_num: int) -> Dict:
        """Xử lý một dòng dữ liệu"""
        processed = row.to_dict()
        error_notes = []

        # Lấy ID Video và Thời gian để tạo link
        video_id = row.get('ID Video', '')
        time_range = row.get('Thời gian', '')

        # Xử lý cột Thời gian (cột H - index 7)
        time_val = row.get('Thời gian', '')
        if pd.notna(time_val):
            formatted_time, duration = self.calculator.parse_time_range(
                str(time_val))
            processed['Thời gian'] = formatted_time
            processed['Thời lượng'] = duration

            # Cập nhật time_range với giá trị đã format
            time_range = formatted_time

            if "error" in formatted_time.lower():
                error_notes.append(
                    f"Lỗi định dạng thời gian ở dòng {excel_row_num}")
        else:
            processed['Thời lượng'] = ""

        # Xử lý ngày tháng và gia hạn
        start_date = row.get('Ngày bắt đầu', '')
        extension_dates = {}  # Dictionary để lưu các ngày gia hạn

        if pd.notna(start_date):
            date_results = self._calculate_dates(start_date)
            processed.update(date_results)
            extension_dates = date_results  # Lưu lại để kiểm tra

        # Tính nhuận bút CƠ BẢN
        usage_type = row.get('Hình thức sử dụng', '')
        base_fee = 0

        if pd.notna(usage_type):
            duration = processed.get('Thời lượng', '')
            share_percent = row.get('Share%', '')

            base_fee, error = self.calculator.calculate_base_fee(
                usage_type, duration, share_percent
            )

            if error:
                error_notes.append(error)

            processed['Mức nhuận bút'] = base_fee

            # QUAN TRỌNG: Chỉ tính mức nhuận bút gia hạn khi CÓ ngày gia hạn tương ứng
            for i in range(1, 6):
                extension_key = f'Gia hạn (lần {i})'
                royalty_key = f'Mức nhuận bút gia hạn (lần {i})'

                # Kiểm tra xem có ngày gia hạn không
                if extension_key in processed:
                    extension_date = processed.get(extension_key, '')

                    # Nếu có ngày gia hạn (không rỗng) thì mới tính mức nhuận bút
                    if extension_date and pd.notna(extension_date) and str(extension_date).strip():
                        # Tính mức nhuận bút gia hạn (40% của mức cơ bản)
                        renewal_fee = int(round(base_fee * 0.4))
                        processed[royalty_key] = renewal_fee
                    else:
                        # Không có ngày gia hạn => không có mức nhuận bút
                        processed[royalty_key] = ''
                else:
                    # Không có cột ngày gia hạn => không có mức nhuận bút
                    processed[royalty_key] = ''
        else:
            # Không có loại hình sử dụng => để trống tất cả mức nhuận bút
            processed['Mức nhuận bút'] = ''
            for i in range(1, 6):
                processed[f'Mức nhuận bút gia hạn (lần {i})'] = ''

        # THÊM MỚI: Tạo link YouTube với timestamp
        processed['Link YouTube Timestamp'] = self._create_youtube_link_with_timestamp(video_id, time_range)

        # Ghi chú lỗi nếu có
        if error_notes:
            existing_error = processed.get('Error', '')
            if existing_error and pd.notna(existing_error):
                processed['Error'] = f"{existing_error}; {'; '.join(error_notes)}"
            else:
                processed['Error'] = '; '.join(error_notes)

        return processed

    def _calculate_dates(self, start_date) -> Dict:
        """Tính toán ngày kết thúc và gia hạn"""
        dates = {}

        start_dt = parse_date(str(start_date))
        if not start_dt:
            return dates

        # Thời hạn kết thúc: +2 năm - 1 ngày
        end_dt = start_dt + relativedelta(years=2) - relativedelta(days=1)
        dates['Thời hạn kết thúc'] = to_ddmmyyyy(end_dt)

        # Tính 5 lần gia hạn - CHỈ khi đã qua thời hạn
        current_date = datetime.now()
        last_date = end_dt

        for i in range(1, 6):
            # Kiểm tra xem có cần gia hạn không (đã qua thời hạn)
            if last_date < current_date:
                # Ngày bắt đầu gia hạn = ngày sau khi kết thúc
                extension_start = last_date + relativedelta(days=1)
                # Ngày kết thúc gia hạn = +2 năm từ ngày bắt đầu gia hạn - 1 ngày
                extension_end = extension_start + relativedelta(years=2) - relativedelta(days=1)
                dates[f'Gia hạn (lần {i})'] = to_ddmmyyyy(extension_end)
                last_date = extension_end
            else:
                # Chưa hết hạn => không cần gia hạn
                dates[f'Gia hạn (lần {i})'] = ''

        return dates

    def _write_formatted_excel(self, df: pd.DataFrame,
                               output_path: str) -> bool:
        """Ghi file Excel với định dạng"""
        try:
            # Ghi DataFrame ra Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Kết quả', index=False)

            # Mở lại để định dạng
            wb = load_workbook(output_path)
            ws = wb['Kết quả']

            # Định dạng font Times New Roman
            font = Font(name='Times New Roman', size=12)
            hyperlink_font = Font(
                name='Times New Roman',
                size=12,
                color='0000FF',
                underline='single')

            # Border
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Alignment giữa
            left_align = Alignment(horizontal='left', vertical='center')

            # Màu vàng cho header và các cột nhuận bút
            yellow_fill = PatternFill(
                start_color="FFFF00",
                end_color="FFFF00",
                fill_type="solid")

            # Định dạng header - màu vàng
            for cell in ws[1]:
                cell.font = font
                cell.border = thin_border
                cell.alignment = left_align
                cell.fill = yellow_fill

            # Tìm vị trí cột Link YouTube Timestamp
            link_col_index = None
            for i, cell in enumerate(ws[1]):
                if cell.value == 'Link YouTube Timestamp':
                    link_col_index = i + 1
                    break

            # Định dạng toàn bộ sheet
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                for col_idx, cell in enumerate(row, start=1):
                    cell.font = font
                    cell.border = thin_border
                    cell.alignment = left_align

                    # Tô màu vàng cho các ô nhuận bút CÓ GIÁ TRỊ
                    # Cột R (18): Mức nhuận bút
                    # Cột S-W (19-23): Mức nhuận bút gia hạn 1-5
                    if 18 <= col_idx <= 23:
                        if cell.value and str(cell.value).strip() and str(cell.value) != '0':
                            cell.fill = yellow_fill
                            # Định dạng số cho các cột nhuận bút
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0'

                    # Định dạng hyperlink cho cột Link YouTube Timestamp
                    if col_idx == link_col_index:
                        link_value = cell.value
                        if link_value and str(link_value).startswith('https://'):
                            cell.hyperlink = str(link_value)
                            cell.font = hyperlink_font

            # Auto-fit columns
            for column_cells in ws.columns:
                length = max(len(str(cell.value or ''))
                             for cell in column_cells)
                # Giới hạn độ rộng cột Link
                if ws[column_cells[0].column_letter + '1'].value == 'Link YouTube Timestamp':
                    ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 70)
                else:
                    ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

            # Thêm hyperlink cho cột ID Video nếu có
            id_col_index = None
            for i, cell in enumerate(ws[1]):
                if cell.value == 'ID Video':
                    id_col_index = i + 1
                    break

            if id_col_index:
                for row in ws.iter_rows(
                        min_row=2, min_col=id_col_index, max_col=id_col_index):
                    for cell in row:
                        video_id = cell.value
                        if video_id and isinstance(
                                video_id, str) and len(video_id) == 11:
                            cell.hyperlink = f"https://www.youtube.com/watch?v={video_id}"
                            cell.font = hyperlink_font

            wb.save(output_path)

            # Log thông tin để debug
            print(f"File đã được lưu: {output_path}")
            print(f"Đã thêm cột 'Link YouTube Timestamp' ở vị trí cuối cùng")

            return True

        except Exception as e:
            print(f"Lỗi khi ghi file Excel: {e}")
            return False