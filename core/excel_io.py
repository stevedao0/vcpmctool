# vcpmctool/core/excel_io.py (Phiên bản cuối cùng)
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill


def read_input_excel(path: str):
    try:
        df = pd.read_excel(path, engine='openpyxl', dtype=str).fillna("")
        return df, True
    except Exception:
        return pd.DataFrame(), False


def write_output_excel(df: pd.DataFrame, path: str,
                       auto_backup: bool) -> bool:
    try:
        # Ghi dữ liệu thô vào file Excel
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Ket qua", index=False)

        # Mở lại file Excel để định dạng và thêm hyperlink
        wb = load_workbook(path)
        ws = wb["Ket qua"]

        # --- Định dạng cơ bản (giữ nguyên) ---
        font = Font(name='Times New Roman', size=12)
        hyperlink_font = Font(
            name='Times New Roman',
            size=12,
            color="0000FF",
            underline="single")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        yellow_fill = PatternFill(
            start_color="FFFF00",
            end_color="FFFF00",
            fill_type="solid")

        # Định dạng header
        for cell in ws[1]:
            cell.font = font
            cell.border = thin_border
            cell.fill = yellow_fill

        # Định dạng các dòng dữ liệu
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = font
                cell.border = thin_border

        # --- LOGIC THÊM HYPERLINK ---
        # 1. Tìm vị trí cột "ID Video"
        id_video_col_index = None
        for i, cell in enumerate(ws[1]):  # Duyệt qua hàng header
            if cell.value == "ID Video":
                id_video_col_index = i + 1  # openpyxl dùng chỉ số từ 1
                break

        # 2. Thêm hyperlink nếu tìm thấy cột
        if id_video_col_index:
            for row in ws.iter_rows(
                    min_row=2,
                    min_col=id_video_col_index,
                    max_col=id_video_col_index):
                for cell in row:
                    video_id = cell.value
                    # Kiểm tra ID hợp lệ (là chuỗi và dài 11 ký tự)
                    if video_id and isinstance(
                            video_id, str) and len(video_id) == 11:
                        cell.hyperlink = f"https://www.youtube.com/watch?v={video_id}"
                        cell.font = hyperlink_font  # Áp dụng font màu xanh cho hyperlink

        wb.save(path)
        return True
    except Exception:
        return False
